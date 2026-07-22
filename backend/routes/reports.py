import io
from flask import Blueprint, request, jsonify, g, send_file
from models import Request as RequestModel, Development
from utils.decorators import token_required
from utils.visibility import get_developer_visible_request_ids

reports_bp = Blueprint("reports", __name__, url_prefix="/api/reports")


def _scoped_and_filtered_query():
    q = RequestModel.query
    role = g.current_role
    if role == "Requester":
        q = q.filter(RequestModel.RequesterID == g.current_user.UserID)
    elif role == "Developer":
        visible_ids = get_developer_visible_request_ids(
            Development.query.all(),
            g.current_user.UserID,
        )
        q = q.filter(RequestModel.RequestID.in_(visible_ids or [-1]))
    # PGC / Admin -> All

    platform_id = request.args.get("platformId")
    status = request.args.get("status")
    priority = request.args.get("priority")
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    if platform_id:
        q = q.filter(RequestModel.PlatformID == platform_id)
    if status:
        q = q.filter(RequestModel.Status == status)
    if priority:
        q = q.filter(RequestModel.Priority == priority)
    if date_from:
        q = q.filter(RequestModel.CreatedDate >= date_from)
    if date_to:
        q = q.filter(RequestModel.CreatedDate <= date_to)
    return q


@reports_bp.route("", methods=["GET"])
@token_required
def get_report():
    """Reports screen: Filters. Returns JSON rows for the on-screen grid."""
    rows = _scoped_and_filtered_query().order_by(RequestModel.CreatedDate.desc()).all()
    return jsonify([r.to_dict() for r in rows])


@reports_bp.route("/export", methods=["GET"])
@token_required
def export_report():
    """Reports screen: Export Excel/PDF. ?format=excel|pdf (default excel)."""
    fmt = request.args.get("format", "excel").lower()
    rows = _scoped_and_filtered_query().order_by(RequestModel.CreatedDate.desc()).all()

    columns = ["RequestID", "Type", "Platform", "Requester", "Priority", "Status", "CreatedDate"]
    data = [[r.RequestID, r.Type, r.platform.PlatformName if r.platform else "",
             r.requester.Name if r.requester else "", r.Priority, r.Status,
             r.CreatedDate.strftime("%Y-%m-%d") if r.CreatedDate else ""] for r in rows]

    if fmt == "pdf":
        return _export_pdf(columns, data)
    return _export_excel(columns, data)


def _export_excel(columns, data):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Requests Report"
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in data:
        ws.append(row)
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(col) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      as_attachment=True, download_name="requests_report.xlsx")


def _export_pdf(columns, data):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    table_data = [columns] + [[str(v) for v in row] for row in data]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    doc.build([table])
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                      download_name="requests_report.pdf")
